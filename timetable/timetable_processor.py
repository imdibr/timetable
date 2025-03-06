import pandas as pd
from openpyxl import Workbook
import random
from datetime import datetime
import os

# Constants
COLUMN_SUBJECT = "Subject Name"
COLUMN_TEACHER = "Teacher Name"
COLUMN_YEAR = "Year Name"
COLUMN_ELECTIVE = "Elective Class Status"
COLUMN_CLASS = "Class Name"

# Define days and periods
days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
periods = ["Period 1", "Period 2", "Period 3", "Period 4", "Period 5", "Period 6"]

# Define sections
sections = [
    "2IT", "2DS", "2IOT", "2CS A", "2CS B", "2CS C", "2CSAIML A", "2CSAIML B", "2CSAIML C",
    "3IT", "3DS", "3IOT", "3CS A", "3CS B", "3CS C", "3CSAIML A", "3CSAIML B", "3CSAIML C"
]

# Global variable to track locked periods
locked_classes = {}

def create_empty_timetable():
    """Creates an empty timetable structure for all sections."""
    return {
        section: {day: {period: "" for period in periods} for day in days_of_week}
        for section in sections
    }

def assign_fixed_classes(timetable, config):
    """
    Assigns fixed classes to the timetable based on configuration.
    
    Args:
        timetable: The timetable dictionary
        config: Dictionary containing all class configurations
    """
    fixed_classes = ["Honours/Minors/Certification Course", "Meeting Hour","Half Day"]
    
    # Assign standard fixed classes
    for section in timetable:
        # Meeting hour for all sections
        timetable[section]["Monday"]["Period 4"] = fixed_classes[1]
        
        # Honors/Minors for all sections
        timetable[section]["Tuesday"]["Period 5"] = fixed_classes[0]
        timetable[section]["Tuesday"]["Period 6"] = fixed_classes[0]
        timetable[section]["Friday"]["Period 5"] = fixed_classes[0]
        timetable[section]["Friday"]["Period 6"] = fixed_classes[0]
        timetable[section]["Saturday"]["Period 5"] = fixed_classes[2]
        timetable[section]["Saturday"]["Period 6"] = fixed_classes[2]

    # Assign HED class for second year sections only
    if config.get('hed_day') and config.get('hed_period'):
        second_year_sections = [section for section in sections if section.startswith("2")]
        hed_day = config['hed_day']
        hed_period = f"Period {config['hed_period']}"
        
        for section in second_year_sections:
            timetable[section][hed_day][hed_period] = "HED Class"
            
        # Lock this period
        if hed_day not in locked_classes:
            locked_classes[hed_day] = []
        locked_classes[hed_day].append(hed_period)
    
    # Assign program electives for third year only
    if config.get('progelec1_day') and config.get('progelec1_period'):
        third_year_sections = [section for section in sections if section.startswith("3")]
        
        elec1_day = config['progelec1_day']
        elec1_period = f"Period {config['progelec1_period']}"
        elec1_name = config.get('progelec1_name', 'Program Elective 1')
        
        for section in third_year_sections:
            timetable[section][elec1_day][elec1_period] = elec1_name
            
        # Lock this period
        if elec1_day not in locked_classes:
            locked_classes[elec1_day] = []
        locked_classes[elec1_day].append(elec1_period)
    
    if config.get('progelec2_day') and config.get('progelec2_period'):
        third_year_sections = [section for section in sections if section.startswith("3")]
        
        elec2_day = config['progelec2_day']
        elec2_period = f"Period {config['progelec2_period']}"
        elec2_name = config.get('progelec2_name', 'Program Elective 2')
        
        for section in third_year_sections:
            timetable[section][elec2_day][elec2_period] = elec2_name
            
        # Lock this period
        if elec2_day not in locked_classes:
            locked_classes[elec2_day] = []
        locked_classes[elec2_day].append(elec2_period)
    
    # Assign lab period for third year
    if config.get('lab_day') and config.get('lab_period'):
        third_year_sections = [section for section in sections if section.startswith("3")]
        
        lab_day = config['lab_day']
        lab_period_start = int(config['lab_period'])
        lab_period_end = lab_period_start + 1  # 2 consecutive periods
        
        # Check if end period is valid
        if lab_period_end <= 6:
            for section in third_year_sections:
                for p in range(lab_period_start, lab_period_end + 1):
                    timetable[section][lab_day][f"Period {p}"] = "Lab Period"
                    
            # Lock these periods
            if lab_day not in locked_classes:
                locked_classes[lab_day] = []
            for p in range(lab_period_start, lab_period_end + 1):
                locked_classes[lab_day].append(f"Period {p}")
    
    # Semester specific electives
    semester_type = config.get('semester_type')
    
    if semester_type == 'even':
        # Global elective (2 consecutive periods)
        if config.get('global_elec_day') and config.get('global_elec_period'):
            third_year_sections = [section for section in sections if section.startswith("3")]
            
            ge_day = config['global_elec_day']
            ge_period_start = int(config['global_elec_period'])
            ge_period_end = ge_period_start + 1  # 2 consecutive periods
            
            # Check if end period is valid
            if ge_period_end <= 6:
                for section in third_year_sections:
                    for p in range(ge_period_start, ge_period_end + 1):
                        timetable[section][ge_day][f"Period {p}"] = "Global Elective"
                        
                # Lock these periods
                if ge_day not in locked_classes:
                    locked_classes[ge_day] = []
                for p in range(ge_period_start, ge_period_end + 1):
                    locked_classes[ge_day].append(f"Period {p}")
                    
    elif semester_type == 'odd':
        # Open electives (3 individual periods)
        for i in range(1, 4):  # For open electives 1, 2, 3
            day_key = f'open_elec{i}_day'
            period_key = f'open_elec{i}_period'
            
            if config.get(day_key) and config.get(period_key):
                third_year_sections = [section for section in sections if section.startswith("3")]
                
                oe_day = config[day_key]
                oe_period = f"Period {config[period_key]}"
                
                for section in third_year_sections:
                    timetable[section][oe_day][oe_period] = f"Open Elective {i}"
                    
                # Lock this period
                if oe_day not in locked_classes:
                    locked_classes[oe_day] = []
                locked_classes[oe_day].append(oe_period)

def assign_lab_sessions(timetable, lab_classes):
    """Assigns lab sessions to available slots."""
    random.shuffle(lab_classes)  # Randomize lab allocation order
    assigned_teachers = {day: set() for day in days_of_week}  # Track assigned teachers per day

    for row in lab_classes:
        class_name = row[COLUMN_CLASS]
        subject = row[COLUMN_SUBJECT]
        teacher_name = row[COLUMN_TEACHER]
        elective_status = row.get(COLUMN_ELECTIVE, "No").strip().lower()  # Handle missing electives safely

        # Skip if class_name is not in the timetable
        if class_name not in timetable:
            print(f"Warning: Class '{class_name}' not found in timetable. Skipping.")
            continue

        # Skip elective labs for second-year students
        if class_name.startswith("2") and elective_status == "yes":
            continue

        # Find available lab slots (two consecutive periods)
        available_slots = [
            (day, periods[i], periods[i + 1])
            for day in days_of_week
            for i in range(len(periods) - 1)
            if timetable[class_name][day][periods[i]] == ""
            and timetable[class_name][day][periods[i + 1]] == ""
            and periods[i] not in locked_classes.get(day, [])
            and periods[i+1] not in locked_classes.get(day, [])
            and teacher_name not in assigned_teachers[day]  # Prevent teacher overlap
        ]
        
        random.shuffle(available_slots)  # Randomize slot selection

        # Assign the lab session if an available slot is found
        for day, period1, period2 in available_slots:
            timetable[class_name][day][period1] = f"{subject} (Lab) - {teacher_name}"
            timetable[class_name][day][period2] = f"{subject} (Lab) - {teacher_name}"
            assigned_teachers[day].add(teacher_name)
            break  # Move to the next lab

def generate_timetable(lab_classes, config):
    """Generates the complete timetable with all assignments."""
    # Reset locked_classes for a fresh generation
    global locked_classes
    locked_classes = {}
    
    timetable = create_empty_timetable()
    assign_fixed_classes(timetable, config)
    assign_lab_sessions(timetable, lab_classes)
    return timetable

def save_timetable_to_excel(timetable, output_dir=None):
    """
    Saves the timetable to Excel files.
    
    Args:
        timetable: The timetable dictionary
        output_dir: Optional directory to save the files
    
    Returns:
        tuple: Paths to the second year and third year timetable files
    """
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Create workbooks for second and third year
    second_year_wb = Workbook()
    third_year_wb = Workbook()
    
    # Process sections by year
    for section in sections:
        # Determine which workbook to use
        if section.startswith("2"):
            wb = second_year_wb
        else:
            wb = third_year_wb
        
        # Remove any invalid characters from section name
        sheet_name = ''.join(c for c in section if c.isalnum() or c.isspace())
        ws = wb.create_sheet(title=sheet_name)
        
        # Add headers
        ws.cell(row=1, column=1, value="Day/Period")
        for col, period in enumerate(periods, start=2):
            ws.cell(row=1, column=col, value=period)
        
        # Add data
        for row, day in enumerate(days_of_week, start=2):
            ws.cell(row=row, column=1, value=day)
            for col, period in enumerate(periods, start=2):
                ws.cell(row=row, column=col, value=timetable[section][day][period])
    
    # Remove the default sheets
    if "Sheet" in second_year_wb.sheetnames:
        second_year_wb.remove(second_year_wb["Sheet"])
    if "Sheet" in third_year_wb.sheetnames:
        third_year_wb.remove(third_year_wb["Sheet"])
    
    # Save the files
    second_year_filename = os.path.join(output_dir or "", f"second_year_timetable_{timestamp}.xlsx")
    third_year_filename = os.path.join(output_dir or "", f"third_year_timetable_{timestamp}.xlsx")
    
    second_year_wb.save(second_year_filename)
    third_year_wb.save(third_year_filename)
    
    return second_year_filename, third_year_filename

def load_lab_classes(file_path):
    """
    Load lab classes from an Excel file.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        list: List of dictionaries containing lab class information
    """
    try:
        df = pd.read_excel(file_path)
        return df.to_dict('records')
    except Exception as e:
        raise Exception(f"Error loading file: {str(e)}")